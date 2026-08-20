#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
import zipfile
from collections import Counter
from collections.abc import Iterable, Iterator
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook  # type: ignore[import-untyped]
except ImportError as exc:  # pragma: no cover - exercised by the operator environment
    raise SystemExit(
        "OPENPYXL_REQUIRED: run with the Codex bundled spreadsheet Python runtime"
    ) from exc


PACKAGE_FORMAT = "yobi-external-catalog-v1"
DATA_ORIGIN = "YOGIYO_PUBLIC_WEB"
SOURCE_PLATFORM = "YOGIYO"
NORMALIZATION_CODE = "REQUIRED_SINGLE_SELECT_ZERO_LIMIT"
SERVICE_AREA_ID = "area_myeongdong"
JSONL_FILES = (
    "merchant.jsonl",
    "menu.jsonl",
    "menu_option_group.jsonl",
    "menu_option_item.jsonl",
    "merchant_source_detail.jsonl",
    "menu_source_detail.jsonl",
    "menu_source_section.jsonl",
    "menu_source_section_item.jsonl",
    "source_option.jsonl",
    "option_group_source_detail.jsonl",
    "catalog_source_payload.jsonl",
)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha256_json(value: Any) -> str:
    return hashlib.sha256(canonical_json(value).encode("utf-8")).hexdigest()


def blank_to_none(value: Any) -> Any:
    if isinstance(value, str) and not value.strip():
        return None
    return value


def workbook_rows(workbook: Any, sheet_name: str) -> Iterator[dict[str, Any]]:
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    for values in rows:
        yield {
            header: blank_to_none(values[index] if index < len(values) else None)
            for index, header in enumerate(headers)
        }


def bool_int(value: Any) -> int:
    return int(value is True)


def dict_or_empty(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def list_or_empty(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def source_id(key: str) -> str:
    return key.rsplit("_", 1)[-1]


def slug(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return normalized or "catalog"


def xlsx_core_rows(
    xlsx_path: Path,
    selected_shop_ids: set[int],
    collected_at: str,
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, dict[str, Any]]]:
    selected_merchant_keys = {f"yogiyo_{shop_id}" for shop_id in selected_shop_ids}
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    tables: dict[str, list[dict[str, Any]]] = {
        "merchant": [],
        "menu": [],
        "menu_option_group": [],
        "menu_option_item": [],
    }
    lookup: dict[str, dict[str, Any]] = {}

    for row in workbook_rows(workbook, "merchants"):
        key = str(row["source_merchant_key"])
        if key not in selected_merchant_keys:
            continue
        if row["data_origin"] != DATA_ORIGIN:
            raise RuntimeError(f"UNEXPECTED_DATA_ORIGIN:{key}")
        normalized = {
            "merchant_id": key,
            "service_area": str(row["service_area"]),
            "service_area_id": SERVICE_AREA_ID,
            "name_ko": str(row["name_ko"]),
            "name_en": row["name_en"],
            "description": row["description"],
            "delivery_fee": int(row["delivery_fee_krw"]),
            "eta_min": int(row["eta_min"]),
            "eta_max": int(row["eta_max"]),
            "min_order_amount": int(row["min_order_amount_krw"]),
            "flavor_profile": row["flavor_profile"],
            "packaging_signal": row["packaging_signal"],
            "is_synthetic": 0,
            "data_origin": DATA_ORIGIN,
            "source_platform": SOURCE_PLATFORM,
            "source_merchant_id": source_id(key),
            "source_collected_at": collected_at,
        }
        tables["merchant"].append(normalized)

    selected_menu_keys: set[str] = set()
    for row in workbook_rows(workbook, "menus"):
        merchant_id = str(row["merchant_ref"])
        if merchant_id not in selected_merchant_keys:
            continue
        key = str(row["source_menu_key"])
        if row["data_origin"] != DATA_ORIGIN:
            raise RuntimeError(f"UNEXPECTED_DATA_ORIGIN:{key}")
        selected_menu_keys.add(key)
        normalized = {
            "menu_id": key,
            "merchant_id": merchant_id,
            "category": str(row["category_name"]),
            "category_id": None,
            "name_ko": str(row["name_ko"]),
            "name_en": row["name_en"],
            "description": row["description"],
            "cultural_description": row["cultural_description"],
            "price": int(row["price_krw"]),
            "serves_min": row["serves_min"],
            "serves_max": row["serves_max"],
            "spice_level": row["spice_level"],
            "dietary_tags_json": [],
            "allergen_tags_json": [],
            "availability": str(row["availability"]),
            "is_synthetic": 0,
            "updated_at": collected_at,
            "data_origin": DATA_ORIGIN,
            "source_platform": SOURCE_PLATFORM,
            "source_menu_id": source_id(key),
            "source_section_id": None,
            "name_en_status": "NOT_PROVIDED" if row["name_en"] is None else "PROVIDED",
            "cultural_description_status": (
                "NOT_PROVIDED" if row["cultural_description"] is None else "PROVIDED"
            ),
            "serves_status": (
                "NOT_PROVIDED"
                if row["serves_min"] is None or row["serves_max"] is None
                else "PROVIDED"
            ),
            "spice_status": "NOT_PROVIDED" if row["spice_level"] is None else "PROVIDED",
            "dietary_data_status": "NOT_PROVIDED",
        }
        tables["menu"].append(normalized)
        lookup[key] = normalized

    selected_group_keys: set[str] = set()
    for row in workbook_rows(workbook, "option_groups"):
        menu_id = str(row["source_menu_key"])
        if menu_id not in selected_menu_keys:
            continue
        key = str(row["source_group_key"])
        selected_group_keys.add(key)
        original_min = int(row["min_select"])
        original_max = int(row["max_select"])
        required = row["required"] is True
        normalization_code = None
        operational_max = original_max
        if original_max < original_min:
            if not (required and original_min == 1 and original_max == 0):
                raise RuntimeError(f"UNAPPROVED_OPTION_NORMALIZATION:{key}")
            operational_max = 1
            normalization_code = NORMALIZATION_CODE
        normalized = {
            "option_group_id": key,
            "menu_id": menu_id,
            "name_en": row["name_en"],
            "name_ko": str(row["name_ko"]),
            "description": row["description"],
            "required": int(required),
            "min_select": original_min,
            "max_select": operational_max,
            "sort_order": int(row["sort_order"]),
            "source_option_group_id": source_id(key),
            "normalization_code": normalization_code,
            "original_min_select": original_min,
            "original_max_select": original_max,
        }
        tables["menu_option_group"].append(normalized)
        lookup[key] = normalized

    for row in workbook_rows(workbook, "option_items"):
        group_id = str(row["source_group_key"])
        if group_id not in selected_group_keys:
            continue
        key = str(row["source_item_key"])
        normalized = {
            "option_item_id": key,
            "option_group_id": group_id,
            "name_en": row["name_en"],
            "name_ko": str(row["name_ko"]),
            "description": row["description"],
            "price_delta": int(row["price_delta_krw"]),
            "availability": str(row["availability"]),
            "dietary_conflict": row["dietary_conflict"],
            "sort_order": int(row["sort_order"]),
            "source_option_item_key": key,
        }
        tables["menu_option_item"].append(normalized)

    workbook.close()
    return tables, lookup


def source_rows(
    raw_zip_path: Path,
    selected_shop_ids: set[int],
    core_tables: dict[str, list[dict[str, Any]]],
    core_lookup: dict[str, dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    with zipfile.ZipFile(raw_zip_path) as archive:
        shops = json.loads(archive.read("shops_raw.json"))
        menu_document = json.loads(archive.read("menus_raw.json"))
    selected_shops = {
        int(shop["id"]): shop for shop in shops if int(shop["id"]) in selected_shop_ids
    }
    selected_responses = {
        int(response["shop_id"]): response
        for response in menu_document["responses"]
        if int(response["shop_id"]) in selected_shop_ids
    }
    if set(selected_shops) != selected_shop_ids or set(selected_responses) != selected_shop_ids:
        raise RuntimeError("SELECTED_SOURCE_RESPONSE_MISSING")

    merchant_by_id = {row["merchant_id"]: row for row in core_tables["merchant"]}
    menu_by_id = {row["menu_id"]: row for row in core_tables["menu"]}
    detail: dict[str, list[dict[str, Any]]] = {
        "merchant_source_detail": [],
        "menu_source_detail": [],
        "menu_source_section": [],
        "menu_source_section_item": [],
        "source_option": [],
        "option_group_source_detail": [],
        "catalog_source_payload": [],
    }

    for shop_id in sorted(selected_shop_ids):
        merchant_id = f"yogiyo_{shop_id}"
        shop = selected_shops[shop_id]
        if merchant_id not in merchant_by_id:
            raise RuntimeError(f"WORKBOOK_MERCHANT_MISSING:{merchant_id}")
        location = dict_or_empty(shop.get("location"))
        review = dict_or_empty(shop.get("review"))
        open_status = dict_or_empty(shop.get("open_status"))
        detail["merchant_source_detail"].append(
            {
                "merchant_id": merchant_id,
                "latitude": location.get("lat"),
                "longitude": location.get("lng"),
                "distance_m": shop.get("distance"),
                "vertical_type": shop.get("vertical_type"),
                "vertical_sub_type": shop.get("vertical_sub_type"),
                "current_open_status": open_status.get("current_open_status"),
                "review_average": review.get("average_rating"),
                "review_count": review.get("count"),
                "review_image_count": review.get("image_count"),
                "review_reply_count": review.get("reply_count"),
                "franchise_json": shop.get("franchise"),
                "vendor_categories_json": list_or_empty(shop.get("vendor_categories")),
                "tags_json": list_or_empty(shop.get("tags")),
                "image_json": dict_or_empty(shop.get("image")),
                "serving_type_json": dict_or_empty(shop.get("serving_type")),
                "representative_menus_json": list_or_empty(shop.get("representative_menus")),
                "operational_json": {
                    key: value
                    for key, value in shop.items()
                    if key
                    not in {
                        "id",
                        "name",
                        "location",
                        "review",
                        "franchise",
                        "vendor_categories",
                        "tags",
                        "image",
                        "serving_type",
                        "representative_menus",
                    }
                },
            }
        )
        detail["catalog_source_payload"].append(
            {
                "payload_id": f"shop_{shop_id}",
                "entity_type": "SHOP",
                "source_entity_id": str(shop_id),
                "payload_sha256": sha256_json(shop),
                "raw_payload": shop,
            }
        )

        response_wrapper = selected_responses[shop_id]
        response = dict_or_empty(response_wrapper.get("menu_response"))
        detail["catalog_source_payload"].append(
            {
                "payload_id": f"menu_response_{shop_id}",
                "entity_type": "MENU_RESPONSE",
                "source_entity_id": str(shop_id),
                "payload_sha256": sha256_json(response_wrapper),
                "raw_payload": response_wrapper,
            }
        )

        raw_menus = dict_or_empty(response.get("menu"))
        for raw_menu_key, raw_menu_value in raw_menus.items():
            raw_menu = dict_or_empty(raw_menu_value)
            menu_id = f"yogiyo_{shop_id}_{raw_menu_key}"
            core_menu = menu_by_id.get(menu_id)
            if core_menu is None:
                raise RuntimeError(f"WORKBOOK_MENU_MISSING:{menu_id}")
            section_id = raw_menu.get("section_id")
            core_menu["source_section_id"] = None if section_id is None else str(section_id)
            merchant_detail = detail["merchant_source_detail"][-1]
            semantic_parts = [
                core_menu["name_ko"],
                core_menu["category"],
                core_menu["description"],
                merchant_by_id[merchant_id]["name_ko"],
                " ".join(str(value) for value in merchant_detail["vendor_categories_json"]),
                " ".join(str(value) for value in merchant_detail["tags_json"]),
            ]
            core_menu["semantic_text"] = " ".join(
                str(value).strip() for value in semantic_parts if value is not None and str(value).strip()
            )
            detail["menu_source_detail"].append(
                {
                    "menu_id": menu_id,
                    "source_section_id": core_menu["source_section_id"],
                    "review_count": raw_menu.get("review_count"),
                    "liquor": bool_int(raw_menu.get("liquor")),
                    "is_adult": bool_int(raw_menu.get("is_adult")),
                    "verified_adult": bool_int(raw_menu.get("verified_adult")),
                    "soldout": bool_int(raw_menu.get("soldout")),
                    "stock_amount": raw_menu.get("stock_amount"),
                    "thumbnail_json": dict_or_empty(raw_menu.get("thumbnail")),
                    "badges_json": list_or_empty(raw_menu.get("badges")),
                    "announcement_json": raw_menu.get("announcement"),
                    "price_json": dict_or_empty(raw_menu.get("price")),
                    "point": raw_menu.get("point"),
                    "point_promotions_json": list_or_empty(raw_menu.get("point_promotions")),
                    "operational_json": {
                        key: value
                        for key, value in raw_menu.items()
                        if key
                        not in {
                            "id",
                            "section_id",
                            "name",
                            "description",
                            "thumbnail",
                            "announcement",
                            "verified_adult",
                            "liquor",
                            "is_adult",
                            "soldout",
                            "stock_amount",
                            "badges",
                            "price",
                            "review_count",
                            "option_sections",
                            "point",
                            "point_promotions",
                        }
                    },
                }
            )
            for raw_group_value in list_or_empty(raw_menu.get("option_sections")):
                raw_group = dict_or_empty(raw_group_value)
                group_id = f"{menu_id}_{raw_group['id']}"
                core_group = core_lookup.get(group_id)
                if core_group is None:
                    raise RuntimeError(f"WORKBOOK_OPTION_GROUP_MISSING:{group_id}")
                detail["option_group_source_detail"].append(
                    {
                        "option_group_id": group_id,
                        "source_option_group_id": str(raw_group["id"]),
                        "multiple_limit": raw_group.get("multiple_limit"),
                        "available_quantity": bool_int(raw_group.get("available_quantity")),
                        "available_multiple": bool_int(raw_group.get("available_multiple")),
                        "original_min_select": core_group["original_min_select"],
                        "original_max_select": core_group["original_max_select"],
                        "badges_json": list_or_empty(raw_group.get("badges")),
                        "tooltip_message": raw_group.get("tooltip_message"),
                        "source_json": raw_group,
                    }
                )

        for section_order, raw_section_value in enumerate(
            list_or_empty(response.get("menu_sections")), start=1
        ):
            raw_section = dict_or_empty(raw_section_value)
            source_section_id = str(raw_section["id"])
            section_key = f"yogiyo_{shop_id}_section_{source_section_id}"
            detail["menu_source_section"].append(
                {
                    "source_section_key": section_key,
                    "merchant_id": merchant_id,
                    "source_section_id": source_section_id,
                    "section_type": raw_section.get("type"),
                    "title": raw_section.get("title"),
                    "description": raw_section.get("description"),
                    "liquor": bool_int(raw_section.get("liquor")),
                    "is_adult": bool_int(raw_section.get("is_adult")),
                    "disposable": bool_int(raw_section.get("disposable")),
                    "additional_discounted": bool_int(
                        raw_section.get("additional_discounted")
                    ),
                    "sort_order": section_order,
                }
            )
            for item_order, raw_menu_id in enumerate(
                list_or_empty(raw_section.get("items")), start=1
            ):
                menu_id = f"yogiyo_{shop_id}_{raw_menu_id}"
                if menu_id not in menu_by_id:
                    raise RuntimeError(f"SECTION_MENU_MISSING:{section_key}:{menu_id}")
                detail["menu_source_section_item"].append(
                    {
                        "source_section_key": section_key,
                        "menu_id": menu_id,
                        "sort_order": item_order,
                    }
                )

        for raw_option_key, raw_option_value in dict_or_empty(response.get("option")).items():
            raw_option = dict_or_empty(raw_option_value)
            price = dict_or_empty(raw_option.get("price"))
            detail["source_option"].append(
                {
                    "source_option_key": f"yogiyo_{shop_id}_{raw_option_key}",
                    "merchant_id": merchant_id,
                    "source_option_id": str(raw_option_key),
                    "name_ko": str(raw_option.get("name") or ""),
                    "description": blank_to_none(raw_option.get("description")),
                    "origin_price": price.get("origin_price"),
                    "final_price": price.get("final_price"),
                    "discount_percent": price.get("discount_percent"),
                    "soldout": bool_int(raw_option.get("soldout")),
                    "stock_amount": raw_option.get("stock_amount"),
                    "deposit_json": dict_or_empty(raw_option.get("deposit")),
                    "reusable_packaging": bool_int(raw_option.get("reusable_packaging")),
                    "source_json": raw_option,
                }
            )

    return detail


def inject_import_id(
    tables: Iterable[list[dict[str, Any]]], catalog_import_id: str
) -> None:
    for rows in tables:
        for row in rows:
            row["catalog_import_id"] = catalog_import_id


def validate_tables(
    tables: dict[str, list[dict[str, Any]]], expected: dict[str, int]
) -> dict[str, Any]:
    count_map = {
        "merchant": len(tables["merchant"]),
        "menu": len(tables["menu"]),
        "menu_option_group": len(tables["menu_option_group"]),
        "menu_option_item": len(tables["menu_option_item"]),
        "merchant_source_detail": len(tables["merchant_source_detail"]),
        "menu_source_detail": len(tables["menu_source_detail"]),
        "menu_source_section": len(tables["menu_source_section"]),
        "menu_source_section_item": len(tables["menu_source_section_item"]),
        "source_option": len(tables["source_option"]),
        "option_group_source_detail": len(tables["option_group_source_detail"]),
        "catalog_source_payload": len(tables["catalog_source_payload"]),
    }
    expected_map = {
        "merchant": int(expected["merchant"]),
        "menu": int(expected["menu"]),
        "menu_option_group": int(expected["menu_option_group"]),
        "menu_option_item": int(expected["menu_option_item"]),
        "merchant_source_detail": int(expected["merchant"]),
        "menu_source_detail": int(expected["menu"]),
        "menu_source_section": int(expected["menu_section"]),
        "menu_source_section_item": int(expected["menu_section_item"]),
        "source_option": int(expected["source_unique_option"]),
        "option_group_source_detail": int(expected["menu_option_group"]),
        "catalog_source_payload": int(expected["raw_source_payload"]),
    }
    mismatches = {
        key: {"expected": expected_map[key], "actual": count_map[key]}
        for key in expected_map
        if count_map[key] != expected_map[key]
    }
    if mismatches:
        raise RuntimeError(f"PACKAGE_COUNT_MISMATCH:{canonical_json(mismatches)}")

    keys = {
        "merchant": "merchant_id",
        "menu": "menu_id",
        "menu_option_group": "option_group_id",
        "menu_option_item": "option_item_id",
        "menu_source_section": "source_section_key",
        "source_option": "source_option_key",
        "catalog_source_payload": "payload_id",
    }
    duplicates: dict[str, list[str]] = {}
    for table, key in keys.items():
        counts = Counter(str(row[key]) for row in tables[table])
        repeated = sorted(value for value, count in counts.items() if count > 1)
        if repeated:
            duplicates[table] = repeated[:10]
    if duplicates:
        raise RuntimeError(f"PACKAGE_DUPLICATE_KEYS:{canonical_json(duplicates)}")

    normalization_count = sum(
        row.get("normalization_code") == NORMALIZATION_CODE
        for row in tables["menu_option_group"]
    )
    null_status_counts = {
        "menu_name_en_not_provided": sum(
            row["name_en_status"] == "NOT_PROVIDED" for row in tables["menu"]
        ),
        "menu_spice_not_provided": sum(
            row["spice_status"] == "NOT_PROVIDED" for row in tables["menu"]
        ),
        "menu_serves_not_provided": sum(
            row["serves_status"] == "NOT_PROVIDED" for row in tables["menu"]
        ),
    }
    return {
        "counts": count_map,
        "normalization_count": normalization_count,
        "null_status_counts": null_status_counts,
        "ingredients_rows": 0,
        "certifications_rows": 0,
        "validation_passed": True,
    }


def zip_info(name: str) -> zipfile.ZipInfo:
    info = zipfile.ZipInfo(name, date_time=(2026, 8, 16, 0, 0, 0))
    info.compress_type = zipfile.ZIP_DEFLATED
    info.external_attr = 0o644 << 16
    return info


def write_package(
    output_path: Path,
    manifest: dict[str, Any],
    selection: dict[str, Any],
    tables: dict[str, list[dict[str, Any]]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output_path, "w", compresslevel=9) as archive:
        archive.writestr(zip_info("manifest.json"), canonical_json(manifest) + "\n")
        archive.writestr(zip_info("selection_manifest.json"), canonical_json(selection) + "\n")
        mapping = {
            "merchant.jsonl": "merchant",
            "menu.jsonl": "menu",
            "menu_option_group.jsonl": "menu_option_group",
            "menu_option_item.jsonl": "menu_option_item",
            "merchant_source_detail.jsonl": "merchant_source_detail",
            "menu_source_detail.jsonl": "menu_source_detail",
            "menu_source_section.jsonl": "menu_source_section",
            "menu_source_section_item.jsonl": "menu_source_section_item",
            "source_option.jsonl": "source_option",
            "option_group_source_detail.jsonl": "option_group_source_detail",
            "catalog_source_payload.jsonl": "catalog_source_payload",
        }
        for filename in JSONL_FILES:
            with archive.open(zip_info(filename), "w") as handle:
                for row in tables[mapping[filename]]:
                    handle.write((canonical_json(row) + "\n").encode("utf-8"))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build a validated, streaming import package for the approved Yogiyo selection."
    )
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--raw-zip", required=True, type=Path)
    parser.add_argument("--summary", required=True, type=Path)
    parser.add_argument("--selection", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    selection = json.loads(args.selection.read_text(encoding="utf-8"))
    summary = json.loads(args.summary.read_text(encoding="utf-8"))
    expected_hashes = selection["source_hashes"]
    actual_hashes = {
        "raw_zip_sha256": sha256_file(args.raw_zip),
        "normalized_xlsx_sha256": sha256_file(args.xlsx),
        "collection_summary_sha256": sha256_file(args.summary),
    }
    if actual_hashes != expected_hashes:
        raise RuntimeError(
            f"SOURCE_HASH_MISMATCH:{canonical_json({'expected': expected_hashes, 'actual': actual_hashes})}"
        )
    if summary != selection["source_summary"]:
        raise RuntimeError("COLLECTION_SUMMARY_MISMATCH")
    if summary["data_origin"] != DATA_ORIGIN or int(summary["menu_response_failure_count"]) != 0:
        raise RuntimeError("SOURCE_SUMMARY_NOT_APPROVED")

    selected_shop_ids = {int(row["shop_id"]) for row in selection["selected_merchants"]}
    if len(selected_shop_ids) != int(selection["selection_limit"]):
        raise RuntimeError("SELECTION_COUNT_MISMATCH")
    collected_at = str(summary["collected_at_utc"])
    core_tables, core_lookup = xlsx_core_rows(args.xlsx, selected_shop_ids, collected_at)
    detail_tables = source_rows(
        args.raw_zip, selected_shop_ids, core_tables, core_lookup
    )
    tables = {**core_tables, **detail_tables}

    selection_sha = sha256_file(args.selection)
    collected_date = datetime.fromisoformat(collected_at).astimezone(timezone.utc).strftime("%Y%m%d")
    algorithm = str(selection["algorithm_version"])
    catalog_import_id = f"yogiyo_{collected_date}_{selection_sha[:16]}"
    catalog_release_id = (
        f"yogiyo-public-web:{collected_date}:{slug(algorithm)}:{selection_sha[:12]}"
    )
    inject_import_id(tables.values(), catalog_import_id)
    diagnostics = validate_tables(tables, selection["related_row_counts"])
    expected_counts = diagnostics["counts"]
    manifest = {
        "package_format": PACKAGE_FORMAT,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "catalog_import_id": catalog_import_id,
        "catalog_release_id": catalog_release_id,
        "data_origin": DATA_ORIGIN,
        "source_platform": SOURCE_PLATFORM,
        "source_hashes": actual_hashes,
        "selection_manifest_sha256": selection_sha,
        "selection_manifest_canonical_sha256": sha256_json(selection),
        "selection_algorithm_version": algorithm,
        "collection_location": str(summary["collection_location"]),
        "source_collected_at": collected_at,
        "service_area_id": SERVICE_AREA_ID,
        "service_area": str(summary["service_area_mapped_to"]),
        "selected_merchant_count": len(selected_shop_ids),
        "expected_counts": expected_counts,
        "diagnostics": diagnostics,
    }
    write_package(args.output, manifest, selection, tables)
    result = {
        "output": str(args.output.resolve()),
        "package_sha256": sha256_file(args.output),
        "catalog_import_id": catalog_import_id,
        "catalog_release_id": catalog_release_id,
        "counts": expected_counts,
        "normalization_count": diagnostics["normalization_count"],
        "source_hashes_verified": True,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
